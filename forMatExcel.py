from openpyxl import load_workbook
import datetime
import time
import re

excel=str(input('请输入excel名称(包含后缀)：'))
wb=load_workbook(excel)
ws=wb.active
check=input('请输入转化为√的区域：')
c=ws[check]
for i in tuple(c):
    for each in i:
        # print(each.value)
        if each.value!=None:
            each.value='√'
        # print(each.value)
print('格式化完成！')

check=input('请输入日期格式化的区域：')
c=ws[check]
for i in tuple(c):
    for each in i:
        # print(each.value)
        t=each.value
        i=datetime.datetime.now()
        if t==None:
            continue
        elif re.match('\d{1,2}\.\d{1,2}',t,)!=None:
            timeStruct = time.strptime(str(i.year)+'.'+t, "%Y.%m.%d")
        elif re.match('\d{1,2}-\d{1,2}',t,)!=None:
            timeStruct = time.strptime(str(i.year)+'-'+t, "%Y-%m-%d")
        elif re.match('^\d{4}-\d{1,2}-\d{1,2}',t,)!=None:
            timeStruct = time.strptime(t, "%Y-%m-%d")
        elif re.match('^\d{4}年\d{1,2}月\d{1,2}日',t,)!=None:
            timeStruct = time.strptime(t, "%Y年%m月%d日")
        elif re.match('\d{1,2}月\d{1,2}日',t,)!=None:
            timeStruct = time.strptime(str(i.year)+'年'+t, "%Y年%m月%d日")
        elif re.match('^\d{4}\.\d{1,2}\.\d{1,2}',t,)!=None:
            timeStruct = time.strptime(t, "%Y.%m.%d")
        else:
            continue
        strTime = time.strftime("%Y.%m.%d", timeStruct)
        each.value=strTime
print('格式化完成！')
wb.save(excel)
wb.close()
